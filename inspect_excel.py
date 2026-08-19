import openpyxl
from pprint import pprint

wb = openpyxl.load_workbook('Machine and Equipment Longevity Tracking Tables.xlsx', data_only=True)
for name in ['BLA']:
    ws = wb[name]
    print(f'\n=== Sheet: {name} ===')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    
    # Find header row - first row with 'Division' in col 1
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        if str(ws.cell(r, 1).value or '').strip().lower() == 'division':
            header_row = r
            break
    
    if header_row:
        headers = []
        for c in range(1, ws.max_column + 1):
            headers.append(str(ws.cell(header_row, c).value) if ws.cell(header_row, c).value else f'Col{c}')
        print('Headers:', headers)
        print('\nFirst 10 data rows:')
        for r in range(header_row + 1, min(header_row + 12, ws.max_row + 1)):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                row.append(str(v) if v is not None else '')
            print(f'Row {r}:', ' | '.join(row[:15]))
        print('\nLast 5 rows:')
        for r in range(max(header_row + 1, ws.max_row - 4), ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                row.append(str(v) if v is not None else '')
            print(f'Row {r}:', ' | '.join(row[:15]))
    else:
        print('No header row found')
