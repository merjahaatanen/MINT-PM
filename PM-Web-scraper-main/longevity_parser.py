"""
longevity_parser.py
===================
Reads the Excel workbook "Machine and Equipment Longevity Tracking Tables.xlsx"
from the project root and normalizes every division sheet into a common
dictionary structure used by the MINT division/department Longevity tabs.

The workbook has one sheet per division.  Header rows and column names differ
slightly across sheets, so this module detects the header by the first row that
contains "Division" in column A and maps known column titles to a common set
of field names.
"""

import json
import os
import re
from collections import OrderedDict
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise ImportError("openpyxl is required to parse longevity data") from exc

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(HERE)


def _find_longevity_workbook() -> str | None:
    """Locate the longevity workbook, preferring the project root (where the
    user placed the .xlsx) and falling back to the app folder."""
    candidates = [
        os.path.join(PROJECT_ROOT, "Machine and Equipment Longevity Tracking Tables.xlsx"),
        os.path.join(HERE, "Machine and Equipment Longevity Tracking Tables.xlsx"),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


DEFAULT_PATH = _find_longevity_workbook() or ""

# User edits to longevity rows are persisted here so Excel stays the source of
# truth while the app can still update values (condition, replacement year, etc.).
EDITS_FILE = os.path.join(PROJECT_ROOT, "longevity_edits.json")

# Common output fields, in the order they should appear in tables.
OUTPUT_FIELDS = [
    "division",
    "item_num",
    "department",
    "dept_key",
    "category",
    "machine",
    "location",
    "type",
    "serial_no",
    "asset_num",
    "year_new",
    "condition",
    "service_status",
    "as_of_year_month",
    "estimated_replacement_year",
    "comments",
]

# Map lower-cased header variants -> normalized field name.
# Values are already run through _normalize_header(), so slashes/parentheses are
# replaced with spaces.
HEADER_ALIASES = OrderedDict([
    ("division", ["division"]),
    ("item_num", ["item#", "item #", "item no", "item no.", "item number"]),
    ("department", ["cell line dept", "cell line deptartment", "dept", "department", "workcell"]),
    ("category", ["category", "machine type"]),
    ("machine", ["machine and equipment", "machine equipment", "equipment"]),
    ("location", ["location and or workcenter", "location", "location and or workcentre"]),
    ("type", ["type capex small tool", "type capex smalltool", "type"]),
    ("serial_no", ["serial no", "serial number", "s n"]),
    ("asset_num", ["asset #", "asset number", "asset no"]),
    ("year_new", ["year new", "year"]),
    ("condition", ["condition"]),
    ("service_status", ["service status in out", "service status", "service status in out)", "status"]),
    ("as_of_year_month", ["as of year month", "as of yearmonth", "as of"]),
    ("estimated_replacement_year", ["estimated replacement year", "est replacement year", "estimted replacement year", "replacement year", "ery"]),
    ("comments", ["comments", "comment"]),
])

# Map free-text department names from the Excel to MINT department keys.
_DEPT_KEY_ALIASES = {
    "assembly": "assembly",
    "soap dispenser assembly": "soap",
    "soap & assembly": "soap",
    "soap and assembly": "soap",
    "soap": "soap",
    "toilet": "toilet",
    "toilet partitions": "toilet",
    "general": "general",
    "machine shop": "machine_shop",
    "maintenance": "maintenance",
    "mfg engineering": "mfg_engineering",
    "manufacturing engineering": "mfg_engineering",
    "me": "mfg_engineering",
    "quality assurance": "quality_assurance",
    "qa": "quality_assurance",
    "shipping": "shipping",
    "fabrication": "machine_shop",
    "tube mill": "machine_shop",
}


def _normalize_header(value: Any) -> str:
    """Lower-case and strip punctuation/spaces so aliases match."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^a-z0-9#]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _map_headers(header_row: list[Any]) -> dict[int, str]:
    """Return {column_index: output_field_name} for known headers."""
    mapped: dict[int, str] = {}
    seen: set[str] = set()
    for idx, raw in enumerate(header_row):
        norm = _normalize_header(raw)
        if not norm:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if norm in aliases and field not in seen:
                mapped[idx] = field
                seen.add(field)
                break
    return mapped


def _coerce_year(value: Any) -> str:
    """Return a 4-digit year string, or empty if unknown."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in ("na", "n/a", "none", "-"):
        return ""
    # Excel sometimes returns ints or datetimes; handle both.
    m = re.search(r"(19|20)\d{2}", s)
    if m:
        return m.group(0)
    return s


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Strip Excel's "00:00:00" time artifacts from dates.
    s = re.sub(r"\s*00:00:00\s*$", "", s)
    return s


def _department_key(dept_name: str) -> str:
    """Best-effort mapping from Excel department text to a MINT dept_key."""
    norm = re.sub(r"[^a-z0-9]", " ", dept_name.lower()).strip()
    norm = re.sub(r"\s+", " ", norm)
    if norm in _DEPT_KEY_ALIASES:
        return _DEPT_KEY_ALIASES[norm]
    # Fuzzy: if any alias is a whole-word substring, use it.
    for alias, key in _DEPT_KEY_ALIASES.items():
        if alias in norm:
            return key
    # Default: slugify the name.
    return re.sub(r"[^a-z0-9]+", "_", norm).strip("_")


# Division sheets we know about.  The key is the lower-cased division code used
# in URLs; the value is the exact Excel sheet name.
DIVISION_SHEETS = OrderedDict([
    ("bla", "BLA"),
    ("bmc", "BMC"),
    ("bwc", "BWC"),
    ("bed", "BED"),
    ("thris", "THRIS"),
])


def _find_header_row(ws) -> int | None:
    """Locate the first row whose column A value is 'division' (case-insensitive)."""
    for r in range(1, min(ws.max_row, 50) + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip().lower() == "division":
            return r
    return None


def _parse_sheet(ws, division: str) -> list[dict[str, str]]:
    """Parse one worksheet into normalized rows."""
    header_row_idx = _find_header_row(ws)
    if not header_row_idx:
        return []

    # Excel often reports a hugely inflated max_column (thousands of phantom
    # empty columns). Scan only up to the last column that actually has a header
    # value so we don't read millions of empty cells per row (which makes the
    # first load hang). Fall back to a small sane cap if nothing is found.
    max_col = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row_idx, c).value not in (None, ""):
            max_col = c
    max_col = max_col or min(ws.max_column, 30)

    header = [ws.cell(header_row_idx, c).value for c in range(1, max_col + 1)]
    mapping = _map_headers(header)

    rows: list[dict[str, str]] = []
    blank_streak = 0
    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw = [_clean_cell(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        # Skip entirely blank rows; bail out once we hit a long run of them so an
        # inflated max_row full of phantom empty rows doesn't stall the load.
        if not any(v for v in raw):
            blank_streak += 1
            if blank_streak >= 50:
                break
            continue
        blank_streak = 0
        # Skip rows where the machine/equipment column is empty.
        machine_idx = next((i for i, f in mapping.items() if f == "machine"), None)
        if machine_idx is not None and not raw[machine_idx].strip():
            continue

        row: dict[str, str] = {field: "" for field in OUTPUT_FIELDS}
        row["division"] = division.upper()

        for idx, field in mapping.items():
            if idx >= len(raw):
                continue
            val = raw[idx]
            if field == "year_new":
                val = _coerce_year(val)
            elif field == "estimated_replacement_year":
                val = _coerce_year(val)
            row[field] = val

        # Derive dept_key from the department text.
        row["dept_key"] = _department_key(row.get("department", ""))

        # If item_num is empty, synthesize one from the row number.
        if not row.get("item_num"):
            row["item_num"] = str(r - header_row_idx)

        rows.append(row)
    return rows


_cache: dict[str, list[dict[str, str]]] | None = None


# Fields the UI is allowed to edit. Department/division are derived, not editable.
EDITABLE_FIELDS = {
    "asset_num", "year_new", "condition", "service_status",
    "estimated_replacement_year", "comments",
}


def _load_edits() -> dict[str, dict[str, dict[str, str]]]:
    """Load the edits overlay, returning {division_key: {item_num: {field: value}}}."""
    if not os.path.exists(EDITS_FILE):
        return {}
    try:
        with open(EDITS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except (json.JSONDecodeError, OSError):
        pass
    return {}


def _save_edits(edits: dict[str, dict[str, dict[str, str]]]) -> None:
    """Persist the edits overlay to disk."""
    try:
        with open(EDITS_FILE, "w", encoding="utf-8") as f:
            json.dump(edits, f, indent=2)
    except OSError:
        pass


def _row_edit_key(department: str, machine: str, item_num: str) -> str:
    """Build the composite key edits are stored under. item_num alone is NOT
    reliably unique (blank Item# cells fall back to a per-sheet row number
    that can collide across departments), so department+machine+item_num is
    used instead to make sure an edit never leaks onto a different row."""
    return f"{(department or '').strip()}||{(machine or '').strip()}||{(item_num or '').strip()}"


def _merge_edits(rows: list[dict[str, str]], division_key: str, edits: dict[str, dict[str, dict[str, str]]]) -> None:
    """Apply persisted edits to parsed rows in-place."""
    div_edits = edits.get(division_key, {})
    for row in rows:
        key = _row_edit_key(row.get("department", ""), row.get("machine", ""), str(row.get("item_num", "")))
        item_edits = div_edits.get(key, {})
        if not item_edits:
            continue
        for field, value in item_edits.items():
            if field in OUTPUT_FIELDS:
                row[field] = value


def apply_edit(division_key: str, item_num: str, updates: dict[str, str],
               department: str = "", machine: str = "") -> dict[str, str] | None:
    """Persist an edit overlay for a single row and return the merged row.

    ``updates`` maps field names to new string values. Unknown fields are ignored.
    ``department``/``machine`` disambiguate rows whose Item# collides with
    another row's (see ``_row_edit_key``); when omitted, falls back to
    matching by item_num alone (legacy behavior).
    """
    global _cache
    division_key = (division_key or "").strip().lower()
    item_num = str(item_num or "").strip()
    if not division_key or not item_num:
        return None

    # Only store whitelisted fields and strip to strings.
    clean: dict[str, str] = {}
    for field, value in updates.items():
        if field in EDITABLE_FIELDS:
            if field in ("year_new", "estimated_replacement_year"):
                clean[field] = _coerce_year(value)
            else:
                clean[field] = str(value).strip()

    key = _row_edit_key(department, machine, item_num)

    edits = _load_edits()
    if clean:
        edits.setdefault(division_key, {})[key] = clean
    else:
        # No valid fields supplied; remove any existing edit record.
        edits.get(division_key, {}).pop(key, None)

    _save_edits(edits)

    # Rebuild the cached row so callers get the updated values immediately.
    if _cache is not None and division_key in _cache:
        for row in _cache[division_key]:
            if (str(row.get("item_num", "")) == item_num
                    and (not department or (row.get("department") or "") == department)
                    and (not machine or (row.get("machine") or "") == machine)):
                for field, value in clean.items():
                    row[field] = value
                return dict(row)
    return None


def load_longevity(path: str = DEFAULT_PATH) -> dict[str, list[dict[str, str]]]:
    """Load and normalize every division sheet from the workbook.

    Returns {division_key: [row, ...]} where division_key is lower-case.
    """
    global _cache
    if _cache is not None:
        return _cache

    path = path or _find_longevity_workbook() or ""
    if not path or not os.path.exists(path):
        _cache = {}
        return _cache

    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}
    edits = _load_edits()
    for div_key, sheet_name in DIVISION_SHEETS.items():
        if sheet_name in wb.sheetnames:
            rows = _parse_sheet(wb[sheet_name], div_key)
            _merge_edits(rows, div_key, edits)
            result[div_key] = rows
        else:
            result[div_key] = []
    _cache = result
    return result


def by_division(division_key: str, path: str = DEFAULT_PATH) -> list[dict[str, str]]:
    """All longevity rows for a single division."""
    data = load_longevity(path)
    return list(data.get(division_key.lower(), []))


def by_department(dept_key: str, division_key: str | None = None, path: str = DEFAULT_PATH) -> list[dict[str, str]]:
    """Longevity rows whose normalized dept_key matches the requested key.
    If division_key is supplied, rows are limited to that division first.
    """
    dept_key = (dept_key or "").strip().lower()
    if division_key:
        rows = by_division(division_key, path)
    else:
        rows = [row for div_rows in load_longevity(path).values() for row in div_rows]
    if not dept_key:
        return rows
    return [row for row in rows if row.get("dept_key") == dept_key]


def invalidate_cache() -> None:
    """Force a re-read of the Excel file on the next call."""
    global _cache
    _cache = None
